import requests
import json
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook
import datetime as dt
import argparse

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Process some values')
parser.add_argument('-k', '--keys', nargs='+', help='additional keys')
parser.add_argument('-c', '--colored', action='store_true', help='flag for coloring')
args = parser.parse_args()

# Get today's date in ISO format
_today = dt.date.today()
today_iso = _today.strftime("%Y-%m-%d")

# Read data from 'vehicles.csv' file
with open('vehicles.csv', 'r', encoding='utf-8') as file:
    csv_data = file.read()

# Send a POST request to the API
url = 'http://127.0.0.1:8000/api/upload-vehicles/'
response = requests.post(url, data={'csv_data': csv_data})
data = response.json()

today = datetime.now()

# Function to calculate color code for 'hu' column
def calculate_color_code_hu(hu):
    if hu is None:
        return None

    hu_date = datetime.strptime(hu, "%Y-%m-%d")
    if hu_date > today:
        return None

    date_gap = today - hu_date
    if date_gap.days < 90:
        return "#007500"  # Green
    elif 90 <= date_gap.days < 365:
        return "#FFA500"  # Orange
    else:
        return "#b30000"  # Red

json_data = data['data']

data_dict = json.loads(json_data)

df = pd.DataFrame(data_dict)

# Sort data by 'gruppe' column
data = df.sort_values(by='gruppe')

if args.keys:
    valid_args = data.columns
    for arg in args.keys:
        if arg not in valid_args:
            print(f"Invalid argument: {arg}")
            exit(1)

# Select columns specified by arguments
selected_columns = ['rnr'] + [key for key in args.keys if key != 'colorCode' and key != 'labelIds']
data_selected = data[selected_columns]

color_selected = [key for key in args.keys if key == 'colorCode' or key == 'labelIds']
color_selected = data[color_selected]

header = data_selected.columns.tolist()

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.append(list(data_selected.columns))

fill_green = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
fill_red = PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")

# Iterate through the data and add rows to the worksheet
for _, row in data.iterrows():
    row_data = [row[column] for column in data_selected.columns]
    ws.append(row_data)

    # Calculate color code for cells based on 'hu' column
    color_code_for_cells = calculate_color_code_hu(row['hu'])

    if args.colored:
        if color_code_for_cells == "#007500":
            for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1,
                                    max_col=len(data_selected.columns)):
                for cell in row:
                    cell.fill = fill_green
        elif color_code_for_cells == "#FFA500":
            for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1,
                                    max_col=len(data_selected.columns)):
                for cell in row:
                    cell.fill = fill_orange
        elif color_code_for_cells == "#b30000":
            for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1,
                                    max_col=len(data_selected.columns)):
                for cell in row:
                    cell.fill = fill_red

# Apply font color to cells in 'labelIds' column
if 'labelIds' in args.keys:
    color_cod_col = color_selected['colorCode']

    for idx in range(len(color_cod_col)):
        color_code = color_cod_col.iloc[idx]
        if not pd.isna(color_code):
            row = ws[idx + 2]  # Adjust the row index
            for cell in row:
                cell.font = Font(color=color_code)

# Save the Excel file with a filename based on today's date
filename = f'vehicles_{today_iso}.xlsx'
wb.save(filename)
